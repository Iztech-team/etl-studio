import csv
import json
import os
from typing import Any, Dict, List, Optional
from utils.sql_parser import SQLParser
from utils.encoding import (
    detect_and_convert,
    detect_encoding,
    normalize_arabic_digits,
    strip_directional_marks,
)
from utils.audit import AuditTrail


class Extractor:
    """Reads xlsx, csv, and SQL dump files from a session directory.

    CSV files stream row-by-row to JSONL files in a staging directory —
    rows are NOT held in `_raw_tables` to keep peak RSS proportional to
    one row at a time, not the whole file. Excel and SQL extraction
    still buffer in memory (those tend to be smaller, and SQL dumps
    need the whole text for parsing anyway).
    """

    def __init__(self, session_dir: str, audit_trail: Optional[AuditTrail] = None):
        self.session_dir = session_dir
        self.audit_trail = audit_trail or AuditTrail()
        self._raw_tables: Dict[str, List[Dict]] = {}
        self._schema: Dict[str, Any] = {}
        self._stats: Dict[str, Any] = {}
        self._previews: Dict[str, List[Dict]] = {}
        self._staging_dir = os.path.join(session_dir, "_staged_rows")
        os.makedirs(self._staging_dir, exist_ok=True)
        self._streamed_tables: Dict[str, str] = {}  # table_name → jsonl path

    # ------------------------------------------------------------------
    def extract_all(self) -> Dict[str, Any]:
        """Synchronous wrapper around extract_all_iter — drains the generator
        and returns the final result. Use the iter version directly when you
        want to surface per-table progress (e.g. streaming resume)."""
        result: Dict[str, Any] = {}
        for event_type, payload in self.extract_all_iter():
            if event_type == "done":
                result = payload
        return result

    def extract_all_iter(self):
        """Generator yielding progress events as files are parsed.

        Events:
          ('start', {'tables': [name, ...], 'total': N})    listdir done
          ('table_done', {                                   one file parsed
              'name': str,
              'rowCount': int,
              'columns': [str, ...],
          })
          ('done', {                                         all done; full result
              'tables', 'schema', 'stats', 'preview', 'ddl_schema'
          })

        SQL files may produce multiple tables per file — we emit a
        'table_done' for each parsed relation so the client sees each
        one tick in.
        """
        files = sorted(os.listdir(self.session_dir))
        # Best-effort up-front list: we know csv/excel filenames map to
        # tables; SQL files are unknown until parsed. Worth surfacing the
        # csv/excel ones immediately so the client can show "X / N".
        upfront_names = []
        for fname in files:
            ext = fname.lower().rsplit(".", 1)[-1]
            if ext == "csv":
                upfront_names.append(fname.rsplit(".", 1)[0])
            elif ext in ("xlsx", "xls"):
                upfront_names.append(fname.rsplit(".", 1)[0])
        yield "start", {
            "tables": upfront_names,
            "total": len(upfront_names),
        }

        for fname in files:
            path = os.path.join(self.session_dir, fname)
            ext = fname.lower().rsplit(".", 1)[-1]
            if ext == "csv":
                self._extract_csv(path, fname)
                table_name = fname.rsplit(".", 1)[0]
                rows = self._raw_tables.get(table_name, [])
                yield "table_done", {
                    "name": table_name,
                    "rowCount": len(rows),
                    "columns": list(rows[0].keys()) if rows else [],
                }
            elif ext in ("xlsx", "xls"):
                # _extract_excel may add multiple sheets as separate tables.
                # Snapshot the table set before/after to figure out which
                # entries it produced, then emit one event per new table.
                before = set(self._raw_tables.keys())
                self._extract_excel(path, fname)
                added = [t for t in self._raw_tables if t not in before]
                for table_name in added:
                    rows = self._raw_tables.get(table_name, [])
                    yield "table_done", {
                        "name": table_name,
                        "rowCount": len(rows),
                        "columns": list(rows[0].keys()) if rows else [],
                    }
            elif ext == "sql":
                before = set(self._raw_tables.keys())
                self._extract_sql(path, fname)
                added = [t for t in self._raw_tables if t not in before]
                for table_name in added:
                    rows = self._raw_tables.get(table_name, [])
                    yield "table_done", {
                        "name": table_name,
                        "rowCount": len(rows),
                        "columns": list(rows[0].keys()) if rows else [],
                    }

        self._infer_schema()
        self._compute_stats()

        # Aggregate per-cell directional-mark / encoding counters into
        # one event per (table, column) triple before downstream code
        # drains the events list.
        self.audit_trail.flush_counters_to_events()

        # Preview prefers the inline-computed sample (CSV-streamed
        # tables) and falls back to slicing rows from `_raw_tables` for
        # Excel/SQL.
        preview = {}
        for table in self._raw_tables:
            inline = self._previews.get(table)
            if inline:
                preview[table] = inline
            else:
                preview[table] = self._raw_tables[table][:5]

        yield "done", {
            "tables": self._raw_tables,
            "schema": self._schema,
            "stats": self._stats,
            "preview": preview,
            "staged_rows": dict(self._streamed_tables),
        }

    # ------------------------------------------------------------------
    def _extract_csv(self, path: str, fname: str):
        """Stream CSV rows row-by-row to a per-table JSONL staging file.

        Schema, stats, and preview are computed inline during the same
        pass so we never need to iterate the rows again. Result: peak
        RSS during extract is proportional to one row at a time, not
        the file size. A 1.1GB CSV that previously needed ~10GB of RAM
        now fits in well under 1GB.

        The staging JSONL is the rows' on-disk representation until
        a downstream consumer (transform, /api/tables/data, etc.) lazy-
        loads them via extract_cache.read_table_rows.
        """
        table_name = fname.rsplit(".", 1)[0]
        encoding = detect_encoding(path)
        staging = self._jsonl_path(table_name)
        preview: List[Dict] = []
        sample: List[Dict] = []  # up to 50 rows used for type inference
        row_count = 0
        columns: List[str] = []
        with open(path, "r", encoding=encoding, errors="replace", newline="") as src, \
             open(staging, "w", encoding="utf-8") as dst:
            reader = csv.DictReader(src)
            for raw in reader:
                cleaned = self._clean_row(raw, table_name)
                if not columns and cleaned:
                    columns = list(cleaned.keys())
                row_count += 1
                if len(preview) < 5:
                    preview.append(cleaned)
                if len(sample) < 50:
                    sample.append(cleaned)
                dst.write(json.dumps(cleaned, ensure_ascii=False))
                dst.write("\n")
        self._streamed_tables[table_name] = staging
        self._raw_tables[table_name] = []  # rows live on disk; consumers lazy-load
        self._previews[table_name] = preview
        self._schema[table_name] = self._infer_columns(sample, columns)
        self._stats[table_name] = {"row_count": row_count}

    def _jsonl_path(self, table_name: str) -> str:
        safe = "".join(c if c.isalnum() or c in "-_." else "_" for c in table_name)
        return os.path.join(self._staging_dir, f"{safe}.jsonl")

    def _clean_row(self, raw: dict, table_name: str) -> dict:
        row = {}
        for k, v in raw.items():
            if isinstance(v, str):
                original = v
                v = strip_directional_marks(v)
                if v != original:
                    self.audit_trail.log_directional_marks_stripped(table_name, k)
            row[k] = v
        return row

    def _infer_columns(self, sample: List[Dict], columns: List[str]) -> Dict:
        if not sample or not columns:
            return {}
        cols = {}
        for col in columns:
            sample_vals = [r[col] for r in sample if r.get(col) not in (None, "")]
            cols[col] = {
                "inferred_type": self._guess_type(sample_vals),
                "nullable": True,
            }
        return cols

    def _extract_excel(self, path: str, fname: str):
        try:
            import openpyxl
        except ImportError:
            return
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [
                str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])
            ]
            data = []
            for row in rows[1:]:
                cleaned_row = {}
                for i in range(len(headers)):
                    v = row[i]
                    if isinstance(v, str):
                        original = v
                        v = strip_directional_marks(v)
                        if v != original:
                            self.audit_trail.log_directional_marks_stripped(
                                sheet, headers[i]
                            )
                    cleaned_row[headers[i]] = v
                data.append(cleaned_row)
            key = (
                f"{fname.rsplit('.', 1)[0]}_{sheet}"
                if len(wb.sheetnames) > 1
                else sheet
            )
            self._raw_tables[key] = data

    def _extract_sql(self, path: str, fname: str):
        content, _enc = detect_and_convert(path)
        parser = SQLParser(content)
        tables = parser.parse()
        self._raw_tables.update(tables)

    # ------------------------------------------------------------------
    def _infer_schema(self):
        # CSV-streamed tables already have their schema computed inline
        # by `_extract_csv`. Only Excel/SQL paths populate `_raw_tables`
        # with actual rows here, so this loop runs over those.
        for table, rows in self._raw_tables.items():
            if table in self._schema:
                continue
            if not rows:
                self._schema[table] = {}
                continue
            cols = {}
            for col in rows[0].keys():
                sample_vals = [
                    r[col] for r in rows[:50] if r.get(col) not in (None, "")
                ]
                cols[col] = {
                    "inferred_type": self._guess_type(sample_vals),
                    "nullable": True,
                }
            self._schema[table] = cols

    def _guess_type(self, vals: list) -> str:
        if not vals:
            return "string"
        int_ok = all(self._is_int(v) for v in vals)
        if int_ok:
            return "integer"
        float_ok = all(self._is_float(v) for v in vals)
        if float_ok:
            return "float"
        bool_ok = all(
            str(v).lower() in ("true", "false", "0", "1", "yes", "no") for v in vals
        )
        if bool_ok:
            return "boolean"
        # date heuristic
        import re

        date_re = re.compile(r"\d{4}-\d{2}-\d{2}")
        if all(date_re.search(str(v)) for v in vals):
            return "date"
        return "string"

    @staticmethod
    def _is_int(v) -> bool:
        try:
            text = normalize_arabic_digits(str(v)).replace(",", "")
            int(text)
            return True
        except Exception:
            return False

    @staticmethod
    def _is_float(v) -> bool:
        try:
            text = normalize_arabic_digits(str(v)).replace(",", "")
            float(text)
            return True
        except Exception:
            return False

    # ------------------------------------------------------------------
    def _compute_stats(self):
        # CSV-streamed tables already have row_count from their
        # streaming pass. Only fill in for Excel/SQL where rows are
        # genuinely in `_raw_tables`.
        for table, rows in self._raw_tables.items():
            if table in self._stats:
                continue
            self._stats[table] = {"row_count": len(rows)}

    # ------------------------------------------------------------------
    def validate(self, config: Dict) -> Dict[str, Any]:
        """Run validation suite and return structured results."""
        issues = []
        record_counts = {t: len(r) for t, r in self._raw_tables.items()}
        financial_totals: Dict[str, Any] = {}
        duplicate_counts: Dict[str, int] = {}
        truncation_risks: list = []
        spot_checks: list = []

        null_values = config.get("null_values", ["", "NULL", "null", "N/A"])

        for table, rows in self._raw_tables.items():
            if not rows:
                issues.append(
                    {
                        "level": "warning",
                        "table": table,
                        "column": None,
                        "message": "Table is empty",
                        "count": 0,
                    }
                )
                continue

            cols = list(rows[0].keys())

            # --- duplicate check (all columns as composite key)
            seen = set()
            dups = 0
            for r in rows:
                key = tuple(strip_directional_marks(str(r.get(c, ""))) for c in cols)
                if key in seen:
                    dups += 1
                seen.add(key)
            duplicate_counts[table] = dups
            if dups:
                issues.append(
                    {
                        "level": "warning",
                        "table": table,
                        "column": None,
                        "message": f"{dups} duplicate row(s) detected",
                        "count": dups,
                    }
                )

            # --- financial totals for numeric columns
            for col in cols:
                vals = [r.get(col) for r in rows]
                num_vals = []
                for v in vals:
                    try:
                        num_vals.append(
                            float(normalize_arabic_digits(str(v)).replace(",", ""))
                        )
                    except Exception:
                        pass
                if len(num_vals) > len(rows) * 0.7:
                    financial_totals[f"{table}.{col}"] = round(sum(num_vals), 4)

                # truncation risk: any string > 255 chars
                str_vals = [str(v) for v in vals if v not in (None, "")]
                long_vals = [v for v in str_vals if len(v) > 255]
                if long_vals:
                    truncation_risks.append(
                        {
                            "table": table,
                            "column": col,
                            "max_length": max(len(v) for v in str_vals),
                            "count": len(long_vals),
                        }
                    )
                    issues.append(
                        {
                            "level": "warning",
                            "table": table,
                            "column": col,
                            "message": f"{len(long_vals)} value(s) exceed 255 chars",
                            "count": len(long_vals),
                        }
                    )

            # --- spot check: first 3 rows
            spot_checks.append({"table": table, "rows": rows[:3]})

        passed = not any(i["level"] == "error" for i in issues)
        return {
            "passed": passed,
            "record_counts": record_counts,
            "financial_totals": financial_totals,
            "duplicate_counts": duplicate_counts,
            "truncation_risks": truncation_risks,
            "issues": issues,
            "spot_checks": spot_checks,
        }
