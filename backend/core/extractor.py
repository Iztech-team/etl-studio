import os
import csv
import json
from typing import Any, Dict, List, Optional
from utils.sql_parser import SQLParser
from utils.encoding import detect_and_convert, normalize_arabic_digits, strip_directional_marks
from utils.audit import AuditTrail


class Extractor:
    """Reads xlsx, csv, and SQL dump files from a session directory."""

    def __init__(self, session_dir: str, audit_trail: Optional[AuditTrail] = None):
        self.session_dir = session_dir
        self.audit_trail = audit_trail or AuditTrail()
        self._raw_tables: Dict[str, List[Dict]] = {}
        self._schema: Dict[str, Any] = {}
        self._stats: Dict[str, Any] = {}
        self._ddl_schema: Dict[str, Any] = {}

    # ------------------------------------------------------------------
    def extract_all(self) -> Dict[str, Any]:
        files = os.listdir(self.session_dir)
        for fname in files:
            path = os.path.join(self.session_dir, fname)
            ext = fname.lower().rsplit(".", 1)[-1]
            if ext == "csv":
                self._extract_csv(path, fname)
            elif ext in ("xlsx", "xls"):
                self._extract_excel(path, fname)
            elif ext == "sql":
                self._extract_sql(path, fname)

        self._infer_schema()
        self._compute_stats()

        return {
            "tables": self._raw_tables,
            "schema": self._schema,
            "stats": self._stats,
            "preview": {t: rows[:5] for t, rows in self._raw_tables.items()},
            "ddl_schema": self._ddl_schema,
        }

    # ------------------------------------------------------------------
    def _extract_csv(self, path: str, fname: str):
        table_name = fname.rsplit(".", 1)[0]
        content, _enc = detect_and_convert(path)
        lines = content.splitlines()
        reader = csv.DictReader(lines)
        cleaned_rows = []
        for r in reader:
            row = {}
            for k, v in r.items():
                if isinstance(v, str):
                    original = v
                    v = strip_directional_marks(v)
                    if v != original:
                        self.audit_trail.log_directional_marks_stripped(table_name, k)
                row[k] = v
            cleaned_rows.append(row)
        self._raw_tables[table_name] = cleaned_rows

    def _extract_excel(self, path: str, fname: str):
        try:
            import openpyxl
        except ImportError:
            return
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [
                str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])
            ]
            data = []
            for row in rows[1:]:
                cleaned_row = {}
                for i in range(len(headers)):
                    v = row[i]
                    if isinstance(v, str):
                        original = v
                        v = strip_directional_marks(v)
                        if v != original:
                            self.audit_trail.log_directional_marks_stripped(sheet, headers[i])
                    cleaned_row[headers[i]] = v
                data.append(cleaned_row)
            key = (
                f"{fname.rsplit('.', 1)[0]}_{sheet}"
                if len(wb.sheetnames) > 1
                else sheet
            )
            self._raw_tables[key] = data

    def _extract_sql(self, path: str, fname: str):
        content, _enc = detect_and_convert(path)
        parser = SQLParser(content)
        tables = parser.parse()
        self._raw_tables.update(tables)
        # Also extract DDL schemas from CREATE TABLE statements
        ddl = parser.parse_ddl()
        # Only keep DDL for tables that have no data rows (DDL-only files)
        for table_name, columns in ddl.items():
            if table_name not in self._raw_tables or not self._raw_tables[table_name]:
                self._ddl_schema[table_name] = columns

    # ------------------------------------------------------------------
    def _infer_schema(self):
        for table, rows in self._raw_tables.items():
            if not rows:
                self._schema[table] = {}
                continue
            cols = {}
            for col in rows[0].keys():
                sample_vals = [
                    r[col] for r in rows[:50] if r.get(col) not in (None, "")
                ]
                cols[col] = {
                    "inferred_type": self._guess_type(sample_vals),
                    "nullable": True,
                }
            self._schema[table] = cols

    def _guess_type(self, vals: list) -> str:
        if not vals:
            return "string"
        int_ok = all(self._is_int(v) for v in vals)
        if int_ok:
            return "integer"
        float_ok = all(self._is_float(v) for v in vals)
        if float_ok:
            return "float"
        bool_ok = all(
            str(v).lower() in ("true", "false", "0", "1", "yes", "no") for v in vals
        )
        if bool_ok:
            return "boolean"
        # date heuristic
        import re

        date_re = re.compile(r"\d{4}-\d{2}-\d{2}")
        if all(date_re.search(str(v)) for v in vals):
            return "date"
        return "string"

    @staticmethod
    def _is_int(v) -> bool:
        try:
            text = normalize_arabic_digits(str(v)).replace(",", "")
            int(text)
            return True
        except Exception:
            return False

    @staticmethod
    def _is_float(v) -> bool:
        try:
            text = normalize_arabic_digits(str(v)).replace(",", "")
            float(text)
            return True
        except Exception:
            return False

    # ------------------------------------------------------------------
    def _compute_stats(self):
        for table, rows in self._raw_tables.items():
            self._stats[table] = {"row_count": len(rows)}

    # ------------------------------------------------------------------
    def validate(self, config: Dict) -> Dict[str, Any]:
        """Run validation suite and return structured results."""
        issues = []
        record_counts = {t: len(r) for t, r in self._raw_tables.items()}
        financial_totals: Dict[str, Any] = {}
        duplicate_counts: Dict[str, int] = {}
        truncation_risks: list = []
        spot_checks: list = []

        null_values = config.get("null_values", ["", "NULL", "null", "N/A"])

        for table, rows in self._raw_tables.items():
            if not rows:
                issues.append(
                    {
                        "level": "warning",
                        "table": table,
                        "column": None,
                        "message": "Table is empty",
                        "count": 0,
                    }
                )
                continue

            cols = list(rows[0].keys())

            # --- duplicate check (all columns as composite key)
            seen = set()
            dups = 0
            for r in rows:
                key = tuple(strip_directional_marks(str(r.get(c, ""))) for c in cols)
                if key in seen:
                    dups += 1
                seen.add(key)
            duplicate_counts[table] = dups
            if dups:
                issues.append(
                    {
                        "level": "warning",
                        "table": table,
                        "column": None,
                        "message": f"{dups} duplicate row(s) detected",
                        "count": dups,
                    }
                )

            # --- financial totals for numeric columns
            for col in cols:
                vals = [r.get(col) for r in rows]
                num_vals = []
                for v in vals:
                    try:
                        num_vals.append(
                            float(normalize_arabic_digits(str(v)).replace(",", ""))
                        )
                    except Exception:
                        pass
                if len(num_vals) > len(rows) * 0.7:
                    financial_totals[f"{table}.{col}"] = round(sum(num_vals), 4)

                # truncation risk: any string > 255 chars
                str_vals = [str(v) for v in vals if v not in (None, "")]
                long_vals = [v for v in str_vals if len(v) > 255]
                if long_vals:
                    truncation_risks.append(
                        {
                            "table": table,
                            "column": col,
                            "max_length": max(len(v) for v in str_vals),
                            "count": len(long_vals),
                        }
                    )
                    issues.append(
                        {
                            "level": "warning",
                            "table": table,
                            "column": col,
                            "message": f"{len(long_vals)} value(s) exceed 255 chars",
                            "count": len(long_vals),
                        }
                    )

            # --- spot check: first 3 rows
            spot_checks.append({"table": table, "rows": rows[:3]})

        passed = not any(i["level"] == "error" for i in issues)
        return {
            "passed": passed,
            "record_counts": record_counts,
            "financial_totals": financial_totals,
            "duplicate_counts": duplicate_counts,
            "truncation_risks": truncation_risks,
            "issues": issues,
            "spot_checks": spot_checks,
        }
